"""
SCB MCP Server — PxWebApi v2
Ger AI-agenter tillgång till SCB:s Statistikdatabas via fem verktyg:
  - scb_search_tables    : Sök bland alla tabeller
  - scb_get_table_info   : Grundinfo om en specifik tabell
  - scb_get_metadata     : Alla variabler och koder för en tabell
  - scb_get_data         : Hämta faktisk statistikdata med filter
  - scb_list_vg_regions  : Alla 49 VG-kommuner med koder (inbyggt)
"""

import asyncio
import itertools
import json
import os

import httpx
from mcp.server.fastmcp import FastMCP
from mcp.server.transport_security import TransportSecuritySettings

# ---------------------------------------------------------------------------
# Konstanter
# ---------------------------------------------------------------------------

SCB_BASE_URL = "https://statistikdatabasen.scb.se/api/v2"
DEFAULT_LANG = "sv"

# FIX Bug B: Separata timeouts — stora datahämtningar behöver mer tid
SCB_META_TIMEOUT  = 30.0   # metadata/sökning
SCB_DATA_TIMEOUT  = 120.0  # datahämtning (stora queries kan ta länge)
KOLADA_TIMEOUT    = 30.0   # per Kolada-anrop

# Västra Götalands alla 49 kommuner med SCB-regionkoder
VG_MUNICIPALITIES: dict[str, str] = {
    "1401": "Härryda",
    "1402": "Partille",
    "1407": "Öckerö",
    "1415": "Stenungsund",
    "1419": "Tjörn",
    "1421": "Orust",
    "1427": "Sotenäs",
    "1430": "Munkedal",
    "1435": "Tanum",
    "1438": "Dals-Ed",
    "1439": "Färgelanda",
    "1440": "Ale",
    "1441": "Lerum",
    "1442": "Vårgårda",
    "1443": "Bollebygd",
    "1444": "Grästorp",
    "1445": "Essunga",
    "1446": "Karlsborg",
    "1447": "Gullspång",
    "1452": "Tranemo",
    "1460": "Bengtsfors",
    "1461": "Mellerud",
    "1462": "Lilla Edet",
    "1463": "Mark",
    "1465": "Svenljunga",
    "1466": "Herrljunga",
    "1470": "Vara",
    "1471": "Götene",
    "1472": "Tibro",
    "1473": "Töreboda",
    "1480": "Göteborg",
    "1481": "Mölndal",
    "1482": "Kungälv",
    "1484": "Lysekil",
    "1485": "Uddevalla",
    "1486": "Strömstad",
    "1487": "Vänersborg",
    "1488": "Trollhättan",
    "1489": "Alingsås",
    "1490": "Borås",
    "1491": "Ulricehamn",
    "1492": "Åmål",
    "1493": "Mariestad",
    "1494": "Lidköping",
    "1495": "Skara",
    "1496": "Skövde",
    "1497": "Hjo",
    "1498": "Tidaholm",
    "1499": "Falköping",
}

VG_COUNTY_CODE = "14"  # Västra Götalands länskod

# ---------------------------------------------------------------------------
# MCP-server
# ---------------------------------------------------------------------------

mcp = FastMCP(
    "scb_mcp",
    transport_security=TransportSecuritySettings(
        enable_dns_rebinding_protection=False
    ),
)

# ---------------------------------------------------------------------------
# Hjälpfunktioner
# ---------------------------------------------------------------------------


async def scb_get(path: str, params: dict | None = None, timeout: float = SCB_META_TIMEOUT) -> dict:
    """Gör ett asynkront GET-anrop mot SCB PxWebApi v2."""
    url = f"{SCB_BASE_URL}/{path}"
    if params is None:
        params = {}
    params.setdefault("lang", DEFAULT_LANG)

    async with httpx.AsyncClient(timeout=timeout) as client:
        try:
            response = await client.get(url, params=params)
            response.raise_for_status()
            return response.json()
        except httpx.HTTPStatusError as e:
            status = e.response.status_code
            messages = {
                400: "Felaktig förfrågan (400): Kontrollera variabelkoder och syntax.",
                403: "Förbjudet (403): Frågan returnerar för många dataceller (max 150 000). Lägg till filter.",
                404: "Hittades inte (404): Kontrollera tabell-ID.",
                429: "För många anrop (429): Vänta några sekunder och försök igen.",
            }
            raise ValueError(messages.get(status, f"API-fel ({status}): {e.response.text[:200]}"))
        except httpx.TimeoutException:
            raise ValueError(f"Timeout: SCB-API:et svarade inte inom {timeout:.0f} sekunder.")


def _format_table_list(tables: list, total: int, page: int, page_size: int) -> str:
    lines = [f"**SCB Tabeller** — {len(tables)} av {total} träffar (sida {page})\n"]
    for t in tables:
        lines.append(f"### {t.get('id', '?')} — {t.get('label', 'Okänd')}")
        lines.append(f"- Uppdaterad: {t.get('updated', '?')[:10]}")
        lines.append(f"- Period: {t.get('firstPeriod', '?')} – {t.get('lastPeriod', '?')}")
        if t.get("variableNames"):
            lines.append(f"- Variabler: {', '.join(t['variableNames'])}")
        lines.append("")
    if total > page * page_size:
        lines.append(f"*Fler sidor tillgängliga. Använd page={page + 1} för nästa sida.*")
    return "\n".join(lines)


def _format_jsonstat2(data: dict, table_id: str) -> str:
    try:
        label = data.get("label", table_id)
        updated = (data.get("updated") or "?")[:10]
        dim_ids: list[str] = data.get("id", [])
        dimensions: dict = data.get("dimension", {})
        values: list = data.get("value", [])

        lines = [f"## {label}", f"*Uppdaterad: {updated}*\n"]

        if not values:
            return "\n".join(lines) + "\nInga datavärden returnerades för valda filter."

        dim_codes: dict[str, list[str]] = {}
        dim_labels: dict[str, dict[str, str]] = {}

        for dim_id in dim_ids:
            dim = dimensions.get(dim_id, {})
            cat = dim.get("category", {})
            idx: dict[str, int] = cat.get("index", {})
            lbl: dict[str, str] = cat.get("label", {})
            sorted_codes = sorted(idx.keys(), key=lambda k: idx[k])
            dim_codes[dim_id] = sorted_codes
            dim_labels[dim_id] = lbl
            preview = ", ".join(lbl.get(c, c) for c in sorted_codes[:8])
            suffix = f" (+{len(sorted_codes) - 8} till)" if len(sorted_codes) > 8 else ""
            lines.append(f"**{dim.get('label', dim_id)}:** {preview}{suffix}")

        lines.append(f"\n**Antal datavärden:** {len(values)}\n")

        if len(values) <= 200:
            lines.append("### Data\n")
            combos = list(itertools.product(*[dim_codes[d] for d in dim_ids]))
            for i, combo in enumerate(combos):
                if i >= len(values):
                    break
                val = values[i]
                parts = [dim_labels[d].get(c, c) for d, c in zip(dim_ids, combo)]
                val_str = str(val) if val is not None else "."
                lines.append(f"- {' | '.join(parts)}: **{val_str}**")
        else:
            lines.append(
                f"*{len(values)} värden — för många att lista. "
                "Använd output_format='json' eller 'csv', eller lägg till filter.*"
            )

        return "\n".join(lines)

    except Exception as exc:
        raw = json.dumps(data, ensure_ascii=False)[:3000]
        return f"Data mottagen men formateringsfel: {exc}\n\nRå data:\n{raw}"


# ---------------------------------------------------------------------------
# Verktyg 1 — Sök tabeller
# ---------------------------------------------------------------------------


@mcp.tool(
    name="scb_search_tables",
    annotations={
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": True,
    },
)
async def scb_search_tables(
    query: str,
    page: int = 1,
    page_size: int = 10,
    only_recent: bool = False,
) -> str:
    """Sök bland alla tabeller i SCB:s Statistikdatabas.

    Returnerar matchande tabeller med ID, titel, uppdateringsdatum, tidsperiod och variabelnamn.
    Nästa steg: anropa scb_get_metadata med table_id för att se variabelkoder.

    Args:
        query: Sökterm på svenska, t.ex. 'befolkning', 'sysselsättning', 'BRP', 'bostäder'
        page: Sidnummer (börjar på 1), default 1
        page_size: Antal resultat per sida (1–50), default 10
        only_recent: True = visa bara tabeller uppdaterade de senaste 30 dagarna
    """
    api_params: dict = {
        "query": query,
        "pageNumber": page,
        "pageSize": max(1, min(50, page_size)),
    }
    if only_recent:
        api_params["pastDays"] = 30

    try:
        data = await scb_get("tables", api_params)
        tables = data.get("tables", [])
        page_info = data.get("page", {})
        total = page_info.get("totalElements", len(tables))
        return _format_table_list(tables, total, page, page_size)
    except ValueError as exc:
        return f"Fel vid sökning: {exc}"


# ---------------------------------------------------------------------------
# Verktyg 2 — Grundinfo om tabell
# ---------------------------------------------------------------------------


@mcp.tool(
    name="scb_get_table_info",
    annotations={
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": True,
    },
)
async def scb_get_table_info(table_id: str) -> str:
    """Hämta grundläggande information om en specifik SCB-tabell.

    Returnerar titel, uppdateringsdatum, tidsperiod, variabelnamn och ämnesväg.

    Args:
        table_id: SCB tabell-ID, t.ex. 'TAB5974' eller 'BE0101A1'
    """
    try:
        data = await scb_get(f"tables/{table_id}")
        lines = [
            f"## Tabell: {data.get('id')} — {data.get('label', '?')}",
            f"- **Uppdaterad:** {(data.get('updated') or '?')[:10]}",
            f"- **Period:** {data.get('firstPeriod', '?')} – {data.get('lastPeriod', '?')}",
            f"- **Källa:** {data.get('source', 'SCB')}",
            f"- **Tidsupplösning:** {data.get('timeUnit', '?')}",
            f"- **Kategori:** {data.get('category', '?')}",
        ]
        if data.get("variableNames"):
            lines.append(f"- **Variabler:** {', '.join(data['variableNames'])}")
        if data.get("paths"):
            path_str = " → ".join(p.get("label", "") for p in data["paths"][0])
            lines.append(f"- **Ämnesväg:** {path_str}")
        lines.append(f"\n*Nästa steg: `scb_get_metadata('{data.get('id')}')` för att se variabelkoder.*")
        return "\n".join(lines)
    except ValueError as exc:
        return f"Fel: {exc}"


# ---------------------------------------------------------------------------
# Verktyg 3 — Metadata / variabelkoder
# ---------------------------------------------------------------------------


@mcp.tool(
    name="scb_get_metadata",
    annotations={
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": True,
    },
)
async def scb_get_metadata(table_id: str) -> str:
    """Hämta detaljerad metadata för en SCB-tabell: variabler, dimensioner och värdekoder.

    Obligatoriskt steg innan datahämtning — du behöver variabelkoderna för filter.
    Visar vilka variabler som är obligatoriska respektive valfria (eliminerbara).

    Args:
        table_id: SCB tabell-ID, t.ex. 'TAB5974'
    """
    try:
        data = await scb_get(f"tables/{table_id}/metadata")
        lines = [f"## Metadata: {data.get('label', table_id)}\n"]
        dimensions: dict = data.get("dimension", {})
        dim_ids: list[str] = data.get("id", [])

        for dim_id in dim_ids:
            dim = dimensions.get(dim_id, {})
            label = dim.get("label", dim_id)
            eliminable = dim.get("extension", {}).get("elimination", False)

            lines.append(f"### `{dim_id}` — {label}")
            lines.append(f"- Obligatorisk: {'Nej (kan utelämnas)' if eliminable else 'Ja (måste anges)'}")

            category = dim.get("category", {})
            codes: dict = category.get("index", {})
            labels: dict = category.get("label", {})

            if codes:
                shown = list(codes.keys())[:25]
                lines.append(f"- Antal värden: {len(codes)}")
                lines.append("- Koder:")
                for code in shown:
                    lines.append(f"  - `{code}` = {labels.get(code, '?')}")
                if len(codes) > 25:
                    lines.append(f"  - *…och {len(codes) - 25} till. Använd `*` för att välja alla.*")
            lines.append("")

        lines += [
            "---",
            "**Användning i scb_get_data:**",
            "- `variable_filters = 'Region=1480,1490;Tid=top(5)'`",
            "- Specialuttryck: `*` (alla), `top(N)`, `from(kod)`, `to(kod)`, `range(kod1,kod2)`",
        ]
        return "\n".join(lines)
    except ValueError as exc:
        return f"Fel: {exc}"


# ---------------------------------------------------------------------------
# Verktyg 4 — Hämta data
# ---------------------------------------------------------------------------


@mcp.tool(
    name="scb_get_data",
    annotations={
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": True,
    },
)
async def scb_get_data(
    table_id: str,
    variable_filters: str = "",
    output_format: str = "readable",
) -> str:
    """Hämta statistikdata från en SCB-tabell med valfria filter.

    Kräver att variabelkoder är kända (kör scb_get_metadata först).
    Stöder kraftfulla filteruttryck: top, from, to, range, wildcard (*).
    Max 150 000 dataceller per anrop — lägg till filter om du får fel 403.

    Args:
        table_id: SCB tabell-ID, t.ex. 'TAB5974'
        variable_filters: Filter på formen 'VarID1=kod1,kod2;VarID2=top(5)'.
            Separera variabler med semikolon och koder med komma.
            Specialuttryck: * (alla), top(N), from(kod), to(kod), range(kod1,kod2).
            Exempel: 'Region=1480,1490;Tid=top(5)' eller 'Region=*;ContentsCode=Folkmängd;Tid=from(2020)'
        output_format: 'readable' (markdown, default), 'json' (JSON-stat2 rådata), 'csv' (semikolon-separerad)
    """
    api_params: dict = {}

    if variable_filters.strip():
        for part in variable_filters.split(";"):
            part = part.strip()
            if "=" in part:
                var_id, codes = part.split("=", 1)
                api_params[f"valueCodes[{var_id.strip()}]"] = codes.strip()

    api_params["outputFormat"] = "json-stat2"

    try:
        # FIX Bug B: använd längre timeout för datahämtning
        data = await scb_get(f"tables/{table_id}/data", api_params, timeout=SCB_DATA_TIMEOUT)

        if output_format == "json":
            return json.dumps(data, ensure_ascii=False, indent=2)
        elif output_format == "csv":
            dim_ids: list[str] = data.get("id", [])
            dimensions: dict = data.get("dimension", {})
            values: list = data.get("value", [])
            dim_codes: dict[str, list[str]] = {}
            dim_labels: dict[str, dict[str, str]] = {}
            for dim_id in dim_ids:
                dim = dimensions.get(dim_id, {})
                cat = dim.get("category", {})
                idx = cat.get("index", {})
                lbl = cat.get("label", {})
                dim_codes[dim_id] = sorted(idx.keys(), key=lambda k: idx[k])
                dim_labels[dim_id] = lbl
            header = ";".join(dim.get("label", d) for d, dim in [(d, dimensions.get(d, {})) for d in dim_ids]) + ";Värde"
            rows = [header]
            combos = list(itertools.product(*[dim_codes[d] for d in dim_ids]))
            for i, combo in enumerate(combos):
                if i >= len(values):
                    break
                parts = [dim_labels[d].get(c, c) for d, c in zip(dim_ids, combo)]
                val_str = str(values[i]) if values[i] is not None else "."
                rows.append(";".join(parts) + f";{val_str}")
            return "\n".join(rows)
        else:
            return _format_jsonstat2(data, table_id)

    except ValueError as exc:
        return f"Fel vid datahämtning: {exc}"


# ---------------------------------------------------------------------------
# Verktyg 5 — Lista VG-regioner
# ---------------------------------------------------------------------------


@mcp.tool(
    name="scb_list_vg_regions",
    annotations={
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
async def scb_list_vg_regions(filter: str = "") -> str:
    """Lista alla 49 kommuner i Västra Götalands län med SCB-regionkoder.

    Dessa koder används som filter i scb_get_data, t.ex. Region=1480,1490.
    Länskoden för hela Västra Götaland är '14'.

    Args:
        filter: Fritext-filter på kommunnamn, t.ex. 'göteborg' eller 'borås' (skiftlägesokänsligt)
    """
    municipalities = dict(VG_MUNICIPALITIES)
    if filter.strip():
        f = filter.lower()
        municipalities = {k: v for k, v in municipalities.items() if f in v.lower()}

    lines = [
        "## Västra Götalands kommuner — SCB-regionkoder",
        f"**Länskod:** `{VG_COUNTY_CODE}` (Västra Götalands län)\n",
        "| Kod | Kommun |",
        "|-----|--------|",
    ]
    for code, name in sorted(municipalities.items(), key=lambda x: x[1]):
        lines.append(f"| `{code}` | {name} |")

    lines.append(f"\n*Totalt: {len(municipalities)} kommuner*")
    lines.append("*Tips: Använd koden i scb_get_data som `Region=1480,1490` i variable_filters.*")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# REST-endpoint för n8n (ingen MCP-sessions-hantering krävs)
# ---------------------------------------------------------------------------

from starlette.requests import Request as StarletteRequest
from starlette.responses import JSONResponse as StarletteJSONResponse
from starlette.routing import Route, Mount
from starlette.applications import Starlette


async def health(request: StarletteRequest):
    """GET /health — Används för att hålla Render.com-instansen varm."""
    return StarletteJSONResponse({"status": "ok"})


async def query_scb(request: StarletteRequest):
    """POST /query — Enkel REST-endpoint för n8n."""
    try:
        body = await request.json()
        table_id = body.get("table_id", "")
        variable_filters = body.get("variable_filters", "")

        if not table_id:
            return StarletteJSONResponse({"error": "table_id krävs"}, status_code=400)

        api_params: dict = {}
        if variable_filters.strip():
            for part in variable_filters.split(";"):
                part = part.strip()
                if "=" in part:
                    var_id, codes = part.split("=", 1)
                    api_params[f"valueCodes[{var_id.strip()}]"] = codes.strip()

        api_params["outputFormat"] = "json-stat2"
        # FIX Bug B: längre timeout för stora datahämtningar
        data = await scb_get(f"tables/{table_id}/data", api_params, timeout=SCB_DATA_TIMEOUT)
        return StarletteJSONResponse(data)

    except ValueError as exc:
        return StarletteJSONResponse({"error": str(exc)}, status_code=500)
    except Exception as exc:
        return StarletteJSONResponse({"error": f"Oväntat fel: {exc}"}, status_code=500)


# ---------------------------------------------------------------------------
# Kolada-endpoint för n8n
# ---------------------------------------------------------------------------

KOLADA_BASE_URL = "https://api.kolada.se/v3"

VG_MUNICIPALITY_IDS = [
    "1401","1402","1407","1415","1419","1421","1427","1430","1435","1438",
    "1439","1440","1441","1442","1443","1444","1445","1446","1447","1452",
    "1460","1461","1462","1463","1465","1466","1470","1471","1472","1473",
    "1480","1481","1482","1484","1485","1486","1487","1488","1489","1490",
    "1491","1492","1493","1494","1495","1496","1497","1498","1499"
]

async def kolada_get(path: str, params: dict) -> dict:
    url = f"{KOLADA_BASE_URL}/{path}"
    async with httpx.AsyncClient(timeout=KOLADA_TIMEOUT, follow_redirects=True) as client:
        response = await client.get(url, params=params)
        response.raise_for_status()
        return response.json()


async def _fetch_kolada_municipality(mun_id: str, kpi_id: str, year) -> list:
    """Hämtar Kolada-data för en enskild kommun. Returnerar tom lista vid fel."""
    params = {
        "kpi_id": kpi_id,
        "municipality_id": mun_id,
        "page_size": 100,
    }
    if year:
        params["year"] = str(year)
    try:
        data = await kolada_get("data", params)
        return data.get("values", [])
    except Exception as exc:
        # FIX Bug D: logga felet istället för att svälja det tyst
        print(f"[kolada] fel för kommun {mun_id}: {exc}")
        return []


async def query_kolada(request: StarletteRequest):
    """POST /kolada — Hämtar Kolada-data för alla VG-kommuner.
    Body: { "kpi_id": "N00708", "year": 2023 (valfritt) }
    """
    try:
        body = await request.json()
        kpi_id = body.get("kpi_id", "")
        year = body.get("year")

        if not kpi_id:
            return StarletteJSONResponse({"error": "kpi_id krävs"}, status_code=400)

        # FIX Bug A: parallella anrop med asyncio.gather istället för sekventiell loop
        tasks = [_fetch_kolada_municipality(mun_id, kpi_id, year) for mun_id in VG_MUNICIPALITY_IDS]
        results = await asyncio.gather(*tasks)
        all_values = [item for sublist in results for item in sublist]

        return StarletteJSONResponse({"values": all_values, "count": len(all_values)})

    except Exception as exc:
        return StarletteJSONResponse({"error": f"Fel: {exc}"}, status_code=500)


# ---------------------------------------------------------------------------
# AF-endpoint för n8n — parsar xlsx från Arbetsförmedlingen
# ---------------------------------------------------------------------------

import io
import re
import openpyxl

AF_TIDSSERIER_URL = "https://arbetsformedlingen.se/statistik/sok-statistik/tidigare-statistik-tidsserier"

VG_KOMMUN_NAMN = set(VG_MUNICIPALITIES.values())


async def query_af(request: StarletteRequest):
    """GET /af — Hämtar AF arbetslöshetsdata för VG-kommuner.

    Query params:
      - kpi: 'arbetslöshet_total' (default) eller 'ungdomsarbetslöshet'
      - period: t.ex. '2026-05' (default: senaste tillgängliga)
      - debug: 'sheets' för att lista fliknamn
    """
    try:
        kpi = request.query_params.get("kpi", "arbetslöshet_total")
        period_filter = request.query_params.get("period", None)

        # Steg 1: Hämta sidan och extrahera xlsx-URL
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            page_resp = await client.get(AF_TIDSSERIER_URL)
            page_resp.raise_for_status()
            html = page_resp.text

        match = re.search(
            r'href="(/download/[^"]+web-inskrivna-arbetslosa-andel-av-bas[^"]+\.xlsx)"',
            html
        )
        if not match:
            return StarletteJSONResponse(
                {"error": "Kunde inte hitta xlsx-URL på AF-sidan"}, status_code=500
            )

        xlsx_url = "https://arbetsformedlingen.se" + match.group(1)

        # Steg 2: Ladda ner xlsx
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            xlsx_resp = await client.get(xlsx_url)
            xlsx_resp.raise_for_status()
            xlsx_bytes = xlsx_resp.content

        # Steg 3: Parsa xlsx med openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

        if request.query_params.get("debug") == "sheets":
            return StarletteJSONResponse({"sheets": wb.sheetnames, "xlsx_url": xlsx_url})

        if request.query_params.get("debug") == "rows":
            ws_debug = wb[wb.sheetnames[2]]  # "Andel"-fliken (index 2)
            rows_debug = list(ws_debug.iter_rows(values_only=True))
            preview = [[str(c) for c in row] for row in rows_debug[:25]]
            return StarletteJSONResponse({"preview": preview})

        # Välj flik — "Andel" innehåller procentdata för alla åldersgrupper
        sheet_name = next(
            (s for s in wb.sheetnames if "andel" in s.lower()),
            wb.sheetnames[0]
        )

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return StarletteJSONResponse({"error": "Tom arbetsbok"}, status_code=500)

        # Steg 4: Hitta rubrikrad — innehåller "PERIOD", "Totalt", "Unga 18-24" etc.
        header_row_idx = None
        period_col_idx = None
        kommun_col_idx = None
        value_col_idx = None

        for i, row in enumerate(rows[:25]):
            row_str = [str(c).strip() if c is not None else "" for c in row]
            if "PERIOD" in row_str:
                header_row_idx = i
                period_col_idx = row_str.index("PERIOD")
                for j, cell in enumerate(row_str):
                    if cell.lower() == "kommun" or "kommun" in cell.lower():
                        kommun_col_idx = j
                        break
                if kpi == "ungdomsarbetslöshet":
                    for j, cell in enumerate(row_str):
                        if "18-24" in cell or "unga" in cell.lower():
                            value_col_idx = j
                            break
                else:
                    for j, cell in enumerate(row_str):
                        if cell.lower() == "totalt":
                            value_col_idx = j
                            break
                break

        if header_row_idx is None:
            return StarletteJSONResponse(
                {"error": "Kunde inte hitta rubrikrad med PERIOD", "sheets": wb.sheetnames},
                status_code=500
            )

        if value_col_idx is None:
            return StarletteJSONResponse(
                {"error": f"Kunde inte hitta värdekolumn för kpi={kpi}"},
                status_code=500
            )

        # Bestäm period — senaste om inget angivet
        if not period_filter:
            latest = ""
            for row in rows[header_row_idx + 1:]:
                if not row or row[period_col_idx] is None:
                    continue
                p = str(row[period_col_idx]).strip()
                if re.match(r"\d{4}-\d{2}", p) and p > latest:
                    latest = p
            period_filter = latest

        # Steg 5: Filtrera VG-kommuner och period, bygg resultat
        results = []
        seen = set()
        for row in rows[header_row_idx + 1:]:
            if not row or row[period_col_idx] is None:
                continue
            period_val = str(row[period_col_idx]).strip()
            if period_val != period_filter:
                continue
            if kommun_col_idx is None or row[kommun_col_idx] is None:
                continue
            kommun_namn = str(row[kommun_col_idx]).strip()
            if kommun_namn not in VG_KOMMUN_NAMN:
                continue
            if kommun_namn in seen:
                continue
            seen.add(kommun_namn)
            value = row[value_col_idx]
            if value is not None:
                try:
                    value = float(value)
                except (TypeError, ValueError):
                    value = None
            kod = next((k for k, v in VG_MUNICIPALITIES.items() if v == kommun_namn), None)
            results.append({
                "kommun_namn": kommun_namn,
                "kommun_kod": kod,
                "period": period_filter,
                "value": value,
                "kpi": kpi,
                "kalla": "AF",
                "enhet": "procent"
            })

        return StarletteJSONResponse({
            "values": results,
            "count": len(results),
            "period": period_filter,
            "sheet": sheet_name,
            "xlsx_url": xlsx_url
        })

    except Exception as exc:
        return StarletteJSONResponse({"error": f"Fel: {exc}"}, status_code=500)


# ---------------------------------------------------------------------------
# Start
# ---------------------------------------------------------------------------

mcp_app = mcp.streamable_http_app()

app = Starlette(routes=[
    Route("/health", endpoint=health, methods=["GET"]),   # FIX Bug C: keep-alive endpoint
    Route("/query", endpoint=query_scb, methods=["POST"]),
    Route("/kolada", endpoint=query_kolada, methods=["POST"]),
    Route("/af", endpoint=query_af, methods=["GET"]),
    Mount("/", app=mcp_app),
])

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
